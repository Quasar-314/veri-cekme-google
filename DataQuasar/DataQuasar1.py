from playwright.sync_api import sync_playwright
from dataclasses import dataclass, asdict, field
import pandas as pd
import os
import sys
import json
import webbrowser
from datetime import datetime, timedelta
from openpyxl import load_workbook
import pywhatkit as pwk
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
    QLabel, QLineEdit, QTableWidget, QTableWidgetItem, QTabWidget, QHeaderView, 
    QTextEdit,  QMessageBox, QComboBox, QFileDialog, QDialog,QListWidget,
    QDialogButtonBox, QCheckBox, QFrame
)
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from PyQt5.QtGui import QIcon, QCursor, QPalette, QColor, QFont,QPixmap
from PyQt5.QtCore import Qt, QSize
from cryptography.fernet import Fernet
import uuid
import socket

# Şifreleme anahtarını dosyadan oku veya oluştur
def get_bireysel_anahtar():
    if os.path.exists("bireysel_anahtar.txt"):
        with open("bireysel_anahtar.txt", "rb") as key_file:
            return key_file.read()
    else:
        key = Fernet.generate_key()
        with open("bireysel_anahtar.txt", "wb") as key_file:
            key_file.write(key)
        return key

KEY = get_bireysel_anahtar()
cipher_suite = Fernet(KEY)
# Giriş şifreleri
VALID_CODES = [
"mX3#z9$1kF!dQ",
"rT7%fL@9x2$Pz",
"W1*pZ9+v2X!5g",
"Q7&mY9#4lR*z1",
"v2*P#x3L!rT8q",
"J1@yL6+o8#mX2",
"Z5#jD3+k4$P@1",
"dF7+qX1*rM!z9",
"k3^X@7$z2!JpY",
"L1*qT9+o8!pX4",
"X9@m#7y3$J*v4",
"f2+P#9k7!T@1v",
"W1^pZ4&d2!kQ9",
"t8+Y@1$j2R%z6",
"J9*qX3+o5!P@2",
"k3&X@4$z1L!t8",
"F7*pY9@2+rT1q",
"W3^dL7!k4#zY2",
"rT8&fL1*+z3@J",
"p5+W#1d3!kR^Y",
"X7*mQ9$2v+P@1",
"J2!dL6+o3#z@5",
"Z1*P#4j9!kT2&",
"Q8^mY5&lR+z9*",
"t2!pX@7*+Jk9#",
"Y3+fR7!q1*dM$",
"k9*W@1+L5$pX7",
"X3&mQ9$2z*o@1",
"W7^fL5*q+Y8@J",
"p1@Q#4z+X!3dT",
"Y9!rT@2+o5*Jk",
"L4*kP3@+z9#dY",
"W5*Q7^j1+rT@2",
"z9!kL2*+X3$mR",
"v2+J9#p@7*X!k",
"J4*QdL1+z6@X8",
"P2@Y7!mT4+kL#",
"fL9#J3+*k2@W1",
"W7+pQ9@2*zX!Y",
"k3^L7#X1*+j@4",
"Z9*pL2&Q4+T@1",
"dF8!Y1*x+P@z3",
"fL6#W9*J2@+pQ",
"X1+z7@d*Q5&Y3",
"p3!L@9+W2*z#f",
"mT6^kQ3@X!7+z",
"Z9+fJ*4@W1&dQ",
"d3*L@5z2!+Qp#",
"Y7&mP9#1+X@3j",
"X8*qP3+Y!k9@L"
]

# Stil tanımlamaları
STYLE_SHEET = """
QMainWindow {
    background-color: #ffffff;
}

QTabWidget::pane {
    border: 1px solid #e4f7f5;
    background: #ffffff;
}

QTabWidget::tab-bar {
    alignment: center;
}

QTabBar::tab {
    background: #ffffff;
    color: #17a2b8;
    padding: 8px 20px;
    border: 1px solid #e4f7f5;
    margin-right: 2px;
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
}

QTabBar::tab:selected {
    background: #17a2b8;
    color: white;
}

QTabBar::tab:hover:!selected {
    background: #13b1c7;
    color: white;
}

QPushButton {
    background-color: #17a2b8;
    color: white;
    border: none;
    padding: 8px 15px;
    border-radius: 4px;
    font-weight: bold;
}

QPushButton:hover {
    background-color: #13b1c7;
    cursor: pointer;
}

QPushButton:disabled {
    background-color: #cccccc;
}

QPushButton#deleteButton {
    background-color: #dc3545;
}

QPushButton#deleteButton:hover {
    background-color: #c82333;
}

QLineEdit {
    padding: 8px;
    border: 1px solid #e4f7f5;
    border-radius: 4px;
    background: white;
}

QLineEdit:focus {
    border: 1px solid #17a2b8;
}

QTextEdit {
    border: 1px solid #e4f7f5;
    border-radius: 4px;
    padding: 5px;
    background: white;
}

QTableWidget {
    background-color: white;
    alternate-background-color: #f8f9fa;
    border: 1px solid #e4f7f5;
    border-radius: 4px;
    gridline-color: #e4f7f5;
}

QTableWidget::item {
    padding: 5px;
}

QTableWidget::item:selected {
    background-color: #17a2b8;
    color: white;
}

QHeaderView::section {
    background-color: #17a2b8;
    color: white;
    padding: 8px;
    border: none;
}

QComboBox {
    padding: 8px;
    border: 1px solid #e4f7f5;
    border-radius: 4px;
    background: white;
}

QComboBox::drop-down {
    border: none;
}

QComboBox::down-arrow {
    image: url(icon/down-arrow.png);
    width: 12px;
    height: 12px;
}

QCheckBox {
    spacing: 8px;
}

QCheckBox::indicator {
    width: 18px;
    height: 18px;
}

QLabel {
    color: #333333;
}

QFrame#separator {
    background-color: #e4f7f5;
}
"""



@dataclass
class İşletme:
    """İşletme verilerini tutar"""
    isim: str = None
    adres: str = None
    website: str = None
    telefon: str = None   
    ortalama_puan: float = None
    enlem: float = None
    boylam: float = None
    

@dataclass
class İşletmeListesi:
    """İşletme nesnelerinin listesini tutar ve Excel'e kaydeder"""
    isletme_listesi: list[İşletme] = field(default_factory=list)
    kayit_yolu = 'cikti'

    def veri_cercevesi(self):
        """işletme_listesini pandas veri çerçevesine dönüştürür"""
        return pd.json_normalize(
            (asdict(isletme) for isletme in self.isletme_listesi), sep="_"
        )

    def excele_kaydet(self, dosya_adi):
        """Veri çerçevesini Excel dosyasına kaydeder"""
        if not os.path.exists(self.kayit_yolu):
            os.makedirs(self.kayit_yolu)
        self.veri_cercevesi().to_excel(f"cikti/{dosya_adi}.xlsx", index=False)



#şifre bölümü



class LoginDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.unique_id = str(uuid.uuid4())
        self.initUI()
        self.check_remembered()  # Uygulama açıldığında dosyayı kontrol et

    def initUI(self):
        self.setWindowTitle("DataQuasar")
        self.setGeometry(100, 100, 350, 250)

        # Ana düzen
        layout = QVBoxLayout()

        # Etiketler ve giriş alanı stili
        self.label = QLabel("Doğrulama Kodunu Girin")
        self.label.setFont(QFont("Arial", 12, QFont.Bold))
        self.label.setStyleSheet("""
            color: #333333;  /* Metin rengi */
            padding: 10px;   /* İç boşluk */
            background-color: #f0f8ff;  /* Açık bir arka plan rengi */
            border: 1px solid #cccccc;  /* Kenarlık */
            border-radius: 5px;  /* Kenar yuvarlama */
            
        """)

        self.code_input = QLineEdit()
        self.code_input.setPlaceholderText("Kodunuzu girin")
        self.code_input.setEchoMode(QLineEdit.Password)
        self.code_input.setStyleSheet("""
            QLineEdit {
                padding: 10px;
                font-size: 12pt;  /* Yazı boyutunu artırdık */
                border: 1px solid #cccccc;
                border-radius: 8px;
                background-color: #f9f9f9;  /* Daha açık bir arka plan rengi */
                color: #333;  /* Yazı rengi */
            }
            QLineEdit:focus {
                border: 1px solid #4CAF50;  /* Odaklanınca sınır rengi */                
            }
            QLineEdit::placeholder {
                color: #999;  /* Placeholder metni rengi */
            }
        """)

        # Beni Hatırla seçeneği
        self.remember_checkbox = QCheckBox("Beni hatırla")
        self.remember_checkbox.setFont(QFont("Arial", 10))
        self.remember_checkbox.setStyleSheet("""
            QCheckBox {
                color: #333;  /* Yazı rengi */                
                padding: 5px;
                border-radius: 5px;  /* Köşeleri yuvarlatma */
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border: 2px solid #dc3545;  /* İşaretleyici kenar rengi */
                border-radius: 3px;  /* İşaretleyici köşeleri yuvarlatma */
                background-color: #f9f9f9;  /* İşaretleyici arka plan rengi */
            }
            QCheckBox::indicator:checked {
                background-color: #e4f7f5;  /* İşaretli durumdaki renk */
                border: 2px solid #dc3545;
                image: url('icon/check.png');  /* Tik işareti simgesi */
            }
        """)

        # Butonlar için stil
        self.submit_button = QPushButton("Giriş")
        self.submit_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                font-weight: bold;
                font-size: 10pt;
                padding: 10px;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #177db8;
            }
        """)
        self.submit_button.setCursor(QCursor(Qt.PointingHandCursor))  # El şeklinde imleç
        

        self.contact_button = QPushButton("İletişim")
        self.contact_button.setStyleSheet("""
            QPushButton {
                color: white; 
                background-color: #dc3545; 
                border: none; 
                padding: 10px 15px; 
                border-radius: 5px; 
                font-size: 18px; 
                
            }
            QPushButton:hover {
                background-color: #c62856;  /* Hover rengi */
            }
        """)
        self.contact_button.setIcon(QIcon("icon/instagram.png"))  # İkon dosyasının yolunu belirtin        
        self.contact_button.setIconSize(QSize(24, 24))  # İkon boyutunu ayarlayın
        self.contact_button.setCursor(QCursor(Qt.PointingHandCursor))  # El şeklinde imleç


        # Düzen bileşenlerini ekle
        layout.addWidget(self.label)
        layout.addWidget(self.code_input)
        layout.addWidget(self.remember_checkbox)
        layout.addWidget(self.submit_button)
        layout.addWidget(self.contact_button)

        self.setLayout(layout)

        # Olay bağlantıları
        self.submit_button.clicked.connect(self.check_code)
        self.contact_button.clicked.connect(self.open_instagram)


    def keyPressEvent(self, event):
        # Kopyalama ve yapıştırma işlemlerini engelle
        if (event.key() == Qt.Key_C and event.modifiers() == Qt.ControlModifier) or \
           (event.key() == Qt.Key_V and event.modifiers() == Qt.ControlModifier):
            return  # Kopyalama ve yapıştırma işlemini engelle
        super().keyPressEvent(event)  # Diğer tuşları normal şekilde işle

    def check_remembered(self):
        try:
            with open("dogrulama_onayı", "rb") as f:  # Dosyayı ikili modda aç
                encrypted_data = f.read()
                decrypted_data = cipher_suite.decrypt(encrypted_data).decode()
                saved_ip, saved_code, saved_remember = decrypted_data.split(',')
                current_ip = self.get_ip_address()
                
                if current_ip == saved_ip and saved_code in VALID_CODES:
                    self.code_input.setText(saved_code)  # Kodu göster
                    self.remember_checkbox.setChecked(saved_remember == "1")  # 'Beni Hatırla' işaretliyse geri yükle
                    self.accept()  # Girişi kabul et
        except (FileNotFoundError, ValueError):
            pass  # Dosya yoksa veya veri yanlışsa hiçbir şey yapma

    def get_ip_address(self):
        # Geçerli IP adresini al
        hostname = socket.gethostname()
        return socket.gethostbyname(hostname)

    def check_code(self):
        code = self.code_input.text().strip()
        current_ip = self.get_ip_address()
        if code in VALID_CODES:
            if self.remember_checkbox.isChecked():
                saved_remember = "1"  # 'Beni Hatırla' seçiliyse '1' olarak kaydedin
            else:
                saved_remember = "0"  # 'Beni Hatırla' seçili değilse '0' olarak kaydedin
            data_to_save = f"{current_ip},{code},{saved_remember}".encode()
            encrypted_data = cipher_suite.encrypt(data_to_save)
            with open("dogrulama_onayı", "wb") as f:  # Dosyayı ikili modda aç
                f.write(encrypted_data)
            self.accept()  # Geçerli kod ise girişi kabul et
        else:
            QMessageBox.warning(self, "Hata", "Geçersiz doğrulama kodu. Lütfen tekrar deneyin.")

    def open_instagram(self):
        # Instagram adresini aç
        webbrowser.open("https://www.instagram.com/3.14quasar/")

# Uygulamayı başlatmak için gereken kod buraya gelecek





class VeriÇekmeThread(QThread):
    sinyal_guncelle = pyqtSignal(list)
    sinyal_tamamlandi = pyqtSignal(list)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.arama_terimleri = []

    def arama_ayarla(self, arama_terimleri):
        self.arama_terimleri = arama_terimleri

    def koordinatlari_ayikla(self, url: str) -> tuple[float, float]:
        """URL'den koordinatları ayıklar"""
        koordinatlar = url.split('/@')[-1].split('/')[0]
        return float(koordinatlar.split(',')[0]), float(koordinatlar.split(',')[1])

    def run(self):
        with sync_playwright() as p:
            tarayici = p.chromium.launch(headless=False)
            sayfa = tarayici.new_page()
            sayfa.goto("https://www.google.com/maps", timeout=60000)
            sayfa.wait_for_timeout(5000)

            tum_isletmeler = []

            for arama_terimi in self.arama_terimleri:
                isletme_listesi = İşletmeListesi()
                sayfa.locator('//input[@id="searchboxinput"]').fill(arama_terimi)
                sayfa.wait_for_timeout(3000)
                sayfa.keyboard.press("Enter")
                sayfa.wait_for_timeout(5000)

                # Kaydırma ve veri toplama işlemi
                sayfa.hover('//a[contains(@href, "https://www.google.com/maps/place")]')
                onceki_sayim = 0

                while True:
                    sayfa.mouse.wheel(0, 10000)
                    sayfa.wait_for_timeout(3000)

                    listings = sayfa.locator('//a[contains(@href, "https://www.google.com/maps/place")]').all()

                    if len(listings) == onceki_sayim:
                        break
                    else:
                        onceki_sayim = len(listings)
                        self.sinyal_guncelle.emit([arama_terimi, str(len(listings))])

                    # Sınır kontrolü
                    if self.veri_sınırı and len(listings) >= self.veri_sınırı:
                        break

                # Belirlenen sınır kadar işletme çek
                for liste in listings[:self.veri_sınırı or len(listings)]:
                    try:
                        liste.click()
                        sayfa.wait_for_timeout(5000)

                        # Title etiketinden ismi al
                        title_text = sayfa.title()
                        isim = title_text.split(' - ')[0] if ' - ' in title_text else title_text

                        adres_xpath = '//button[@data-item-id="address"]//div[contains(@class, "fontBodyMedium")]'
                        website_xpath = '//a[@data-item-id="authority"]//div[contains(@class, "fontBodyMedium")]'
                        telefon_xpath = '//button[contains(@data-item-id, "phone:tel:")]//div[contains(@class, "fontBodyMedium")]'                       
                        puan_xpath = '//div[@jsaction="pane.reviewChart.moreReviews"]//div[@role="img"]'
                        

                        isletme = İşletme()
                        isletme.isim = isim

                        isletme.adres = sayfa.locator(adres_xpath).inner_text() if sayfa.locator(adres_xpath).count() > 0 else ""
                        isletme.website = sayfa.locator(website_xpath).inner_text() if sayfa.locator(website_xpath).count() > 0 else ""
                        isletme.telefon = sayfa.locator(telefon_xpath).inner_text() if sayfa.locator(telefon_xpath).count() > 0 else ""
                        
                        
                        
                        if sayfa.locator(puan_xpath).count() > 0:
                            isletme.ortalama_puan = float(sayfa.locator(puan_xpath).get_attribute('aria-label').split()[0].replace(',','.'))
                        else:
                            isletme.ortalama_puan = 0.0

                        isletme.enlem, isletme.boylam = self.koordinatlari_ayikla(sayfa.url)

                        isletme_listesi.isletme_listesi.append(isletme)
                    except Exception as e:
                        print(f'Hata oluştu: {e}')
                        continue

                # Excel'e kaydet
                isletme_listesi.excele_kaydet(f"{arama_terimi}".replace(' ', '_'))
                tum_isletmeler.extend(isletme_listesi.isletme_listesi)

            tarayici.close()
            self.sinyal_tamamlandi.emit(tum_isletmeler)

class ExcelViewer(QWidget):
    def __init__(self):
        super().__init__()
        self.file_paths = []
        self.file_names = []
        self.country_data = self.load_country_data()
        self.initUI()

    def load_country_data(self):
        try:
            with open('alankodu/ulke.json', 'r', encoding='utf-8') as file:
                return json.load(file)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Ülke verileri yüklenemedi: {e}")
            return []

    def initUI(self):
        # Modern Style with Responsive Layout
        self.setStyleSheet("""
            QWidget {
                background-color: #f4f6f9;
                color: #333333;
                font-family: 'Segoe UI', sans-serif;
            }

            QPushButton {
                padding: 12px 20px;
                border: none;
                border-radius: 10px;
                font-weight: 600;
                font-size: 16px;
                margin: 8px;
                transition: all 0.3s ease;
                cursor: pointer;
                text-align: left;
            }

            QPushButton:hover {
                transform: translateY(-2px);
                box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            }

            QPushButton[class="excel"] {
                background-color: #17a2b8;
                color: white;
            }

            QPushButton[class="whatsapp"] {
                background-color: #25D366;
                color: white;
            }

            QPushButton[class="delete"] {
                background-color: #dc3545;
                color: white;
            }

            QPushButton QIcon {
                width: 24px;  /* Increased icon size */
                height: 24px; /* Increased icon size */
                margin-right: 10px; /* Space between icon and text */
            }

            QFrame {
                background-color: white;
                border-radius: 12px;
                box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                padding: 15px;
                margin: 10px;
                overflow: hidden;
            }

            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 10px;
                background-color: white;
                padding: 5px;
                font-size: 14px;
                color: #333;
            }

            QTableWidget::item {
                padding: 8px;
            }

            QTableWidget::horizontalHeader {
                background-color: #f1f1f1;
                border: none;
                font-weight: bold;
            }

            QTableWidget::verticalHeader {
                background-color: #f1f1f1;
                border: none;
            }

            QComboBox {
                border-radius: 6px;
                padding: 8px;
                background-color: #f8f9fa;
                font-size: 13px;
            }

            QCheckBox {
                font-size: 14px;
            }

            QTextEdit {
                border-radius: 6px;
                padding: 10px;
                background-color: #f8f9fa;
                font-size: 14px;
                color: #333;
                min-height: 100px;
            }

            QListWidget {
                border: 1px solid #ddd;
                border-radius: 6px;
                background-color: #ffffff;
            }

            QScrollBar:vertical {
                border: none;
                background: #f1f1f1;
                width: 8px;
            }

            QScrollBar::handle:vertical {
                background: #ccc;
                border-radius: 4px;
            }

            QScrollBar::handle:vertical:hover {
                background: #bbb;
            }
        """)

        # Main horizontal layout
        main_layout = QHBoxLayout()

        # Left section - 30% width
        left_section = QVBoxLayout()

        # Country Code Section
        country_frame = QFrame()
        country_layout = QVBoxLayout(country_frame)

        self.combo_box = QComboBox()
        self.update_country_combo()
        country_layout.addWidget(QLabel("Alan Kodu Seç:"))
        country_layout.addWidget(self.combo_box)

        country_button_layout = QHBoxLayout()
        edit_button = QPushButton('Düzenle')
        edit_button.setProperty('class', 'excel')
        edit_button.setIcon(QIcon("icon/edit.png"))
        edit_button.clicked.connect(self.openEditDialog)

        add_button = QPushButton('Ekle')
        add_button.setProperty('class', 'excel')
        add_button.setIcon(QIcon("icon/add.png"))
        add_button.clicked.connect(self.addOrUpdateCountryCode)

        country_button_layout.addWidget(edit_button)
        country_button_layout.addWidget(add_button)
        country_layout.addLayout(country_button_layout)
        left_section.addWidget(country_frame)

        # WhatsApp Section
        whatsapp_frame = QFrame()
        whatsapp_layout = QVBoxLayout(whatsapp_frame)

        self.message_input = QTextEdit()
        self.message_input.setPlaceholderText("Mesajınızı buraya yazın...")
        whatsapp_layout.addWidget(self.message_input)

        whatsapp_button_layout = QHBoxLayout()
        qr_button = QPushButton('WhatsApp QR')
        qr_button.setProperty('class', 'whatsapp')
        qr_button.setIcon(QIcon("icon/qrwp.png"))
        qr_button.clicked.connect(self.openWhatsAppWeb)

        send_button = QPushButton('Toplu Mesaj')
        send_button.setProperty('class', 'whatsapp')
        send_button.setIcon(QIcon("icon/whatsapp.png"))
        send_button.clicked.connect(self.sendWhatsAppMessageToAll)

        whatsapp_button_layout.addWidget(qr_button)
        whatsapp_button_layout.addWidget(send_button)

        whatsapp_layout.addLayout(whatsapp_button_layout)

        self.checkbox = QCheckBox("İSİMLER Mesaja Ekle")
        whatsapp_layout.addWidget(self.checkbox)
        left_section.addWidget(whatsapp_frame)

        # Center section - 30% width
        center_section = QVBoxLayout()

        excel_frame = QFrame()
        excel_layout = QVBoxLayout(excel_frame)

        self.excel_button = QPushButton('Excel Dosyalarını Seç')
        self.excel_button.setProperty('class', 'excel')
        self.excel_button.setIcon(QIcon("icon/excel.png"))
        self.excel_button.clicked.connect(self.openFileDialog)
        excel_layout.addWidget(self.excel_button)

        self.file_list_widget = QListWidget()
        self.file_list_widget.setMaximumHeight(100)
        excel_layout.addWidget(self.file_list_widget)

        remove_file_button = QPushButton('Dosya Kaldır')
        remove_file_button.setProperty('class', 'delete')
        remove_file_button.setIcon(QIcon("icon/delete.png"))
        remove_file_button.clicked.connect(self.removeSelectedFile)
        excel_layout.addWidget(remove_file_button)

        center_section.addWidget(excel_frame)

        # Right section - 40% width
        right_section = QVBoxLayout()

        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(2)
        self.table_widget.setHorizontalHeaderLabels(["İsimler", "Numaralar"])
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table_widget.setMinimumHeight(300)
        right_section.addWidget(self.table_widget)

        delete_frame = QFrame()
        delete_layout = QVBoxLayout(delete_frame)

        delete_button = QPushButton('Seçilenleri Sil')
        delete_button.setProperty('class', 'delete')
        delete_button.setIcon(QIcon("icon/delete.png"))
        delete_button.clicked.connect(self.deleteSelected)
        delete_layout.addWidget(delete_button)

        right_section.addWidget(delete_frame)

        # Add sections to main layout
        main_layout.addLayout(left_section, 3)
        main_layout.addLayout(center_section, 3)
        main_layout.addLayout(right_section, 4)

        self.setLayout(main_layout)




    def update_country_combo(self):
        self.combo_box.clear()
        for country in self.country_data:
            self.combo_box.addItem(f"{country['name']} ({country['dial_code']})", country['dial_code'])

    
    def save_country_data(self):
        try:
            with open('alankodu/ulke.json', 'w', encoding='utf-8') as file:
                json.dump(self.country_data, file, ensure_ascii=False, indent=4)
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Ülke verileri kaydedilemedi: {e}")

    def openEditDialog(self):
        """Open dialog to edit country dial code"""
        selected_index = self.combo_box.currentIndex()
        if selected_index == -1:
            QMessageBox.warning(self, "Uyarı", "Hiçbir ülke seçili değil.")
            return

        selected_country = self.country_data[selected_index]
        current_dial_code = selected_country['dial_code']

        dialog = QDialog(self)
        dialog.setWindowTitle(f"{selected_country['name']} İçin Alan Kodunu Düzenle")
        layout = QVBoxLayout(dialog)

        dial_code_input = QLineEdit(current_dial_code)
        layout.addWidget(QLabel(f"{selected_country['name']} için alan kodunu düzenleyin:"))
        layout.addWidget(dial_code_input)

        button_box = QDialogButtonBox()
        save_button = button_box.addButton("Kaydet", QDialogButtonBox.AcceptRole)
        cancel_button = button_box.addButton("İptal", QDialogButtonBox.RejectRole)

        save_button.clicked.connect(lambda: self.save_dial_code(dialog, selected_index, dial_code_input.text()))
        cancel_button.clicked.connect(dialog.reject)
        layout.addWidget(button_box)

        dialog.exec_()

    def save_dial_code(self, dialog, index, new_dial_code):
        """Save the new dial code for a specific country"""
        # Validate dial code (optional, but recommended)
        if not new_dial_code.startswith('+'):
            new_dial_code = '+' + new_dial_code
    
        self.country_data[index]['dial_code'] = new_dial_code
        self.save_country_data()
        self.update_country_combo()
        dialog.accept()

    def openFileDialog(self):
        options = QFileDialog.Options()
        file_paths, _ = QFileDialog.getOpenFileNames(
            self, "Excel Dosyalarını Seç", "",
            "Excel Dosyaları (*.xlsx);;Tüm Dosyalar (*)",
            options=options
        )
        if file_paths:
            for file_path in file_paths:
                file_name = os.path.basename(file_path)  # Dosya adını buradan alıyoruz
        
                # Check if file is already loaded
                if file_path not in self.file_paths:
                    self.file_paths.append(file_path)
                    
                    # Add to file list widget
                    self.file_list_widget.addItem(file_name)
                    
                    # Read Excel file
                    self.readExcel(file_path)
                else:
                    QMessageBox.warning(self, "Uyarı", f"{file_name} zaten yüklenmiş.")
        

    

    def readExcel(self, file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
            name = row[0] if row[0] is not None else ""
            phone_number = row[3] if row[3] is not None else ""

            # Ensure phone number is a string and check if it exists in the table
            if isinstance(phone_number, str):
                phone_exists = False
                total_rows = self.table_widget.rowCount()

                # Check if the phone number already exists in the table
                for row_idx in range(total_rows):
                    existing_phone_item = self.table_widget.item(row_idx, 1)
                    if existing_phone_item and existing_phone_item.text() == phone_number:
                        phone_exists = True
                        break

                if not phone_exists:
                    row_position = self.table_widget.rowCount()
                    self.table_widget.insertRow(row_position)
                    self.table_widget.setItem(row_position, 0, QTableWidgetItem(name))
                    self.table_widget.setItem(row_position, 1, QTableWidgetItem(phone_number))


    def addOrUpdateCountryCode(self):
        selected_country_code = self.combo_box.currentData()
        total_rows = self.table_widget.rowCount()

        for row in range(total_rows):
            phone_number_item = self.table_widget.item(row, 1)
            if phone_number_item:
                original_number = phone_number_item.text()

                # Remove existing country codes
                for code in [country['dial_code'] for country in self.country_data]:
                    if original_number.startswith(code):
                        original_number = original_number[len(code):]
                        break

                # Add new country code
                phone_number_item.setText(selected_country_code + original_number)

        

    def sendWhatsAppMessageToAll(self):
        total_rows = self.table_widget.rowCount()
        message = self.message_input.toPlainText().strip()
        
        if not message:
            QMessageBox.warning(self, "Uyarı", "Gönderilecek mesajı girin.")
            return
        
        if total_rows == 0:
            QMessageBox.warning(self, "Uyarı", "Gönderilecek numara bulunamadı.")
            return
        
        for row in range(total_rows):
            phone_number_item = self.table_widget.item(row, 1)
            name_item = self.table_widget.item(row, 0)
        
            if phone_number_item and name_item:
                phone_number = phone_number_item.text()
                company_name = name_item.text()
        
                # Eğer checkbox işaretliyse, isimleri ekle
                if self.checkbox.isChecked():
                    customized_message = message.replace("İSİMLER", company_name)
                else:
                    customized_message = message  # Checkbox işaretli değilse sadece mesaj gönder
        
                try:
                    now = datetime.now()
                    send_time = now + timedelta(seconds=10 * (row + 1))
        
                    pwk.sendwhatmsg_instantly(
                        phone_number,
                        customized_message,
                        wait_time=20,
                        tab_close=True
                    )
                except pwk.core.exceptions.CountryCodeException:
                    # Hata durumunda kullanıcıya uygun mesaj göster
                    QMessageBox.warning(self, "Hata", f"Alan Kodu Eksik: {phone_number}. Lütfen geçerli bir telefon numarası ile deneyin.")
                    continue  # Hata alındığında, o numara için işlem yapmayı durdurur
                    
        QMessageBox.information(self, "İşlem Tamamlandı", "Tüm mesajlar başarıyla gönderildi.")
        

    def openWhatsAppWeb(self):
        webbrowser.open("https://web.whatsapp.com/")

    def deleteSelected(self):
    # Seçili satırların indekslerini bir set içinde toplar (aynı satır birden fazla kez eklenmez)
        selected_rows = set()
        for item in self.table_widget.selectedItems():
            selected_rows.add(item.row())
        
        # Satırları ters sırayla siler (ilk önce en büyük satırı silmek güvenli)
        for row in sorted(selected_rows, reverse=True):
            self.table_widget.removeRow(row)

    def removeSelectedFile(self):
        # Get current selected item
        current_item = self.file_list_widget.currentItem()
        
        if not current_item:
            QMessageBox.warning(self, "Uyarı", "Kaldırılacak dosya seçin.")
            return
        
        # Get index of selected item
        current_index = self.file_list_widget.row(current_item)
        
        # Remove from file paths
        removed_path = self.file_paths.pop(current_index)
        
        # Remove from list widget
        self.file_list_widget.takeItem(current_index)
        
        # Remove corresponding rows from table
        file_name = os.path.basename(removed_path)
        rows_to_remove = []
        
        # Find rows to remove
        for row in range(self.table_widget.rowCount()):
            name_item = self.table_widget.item(row, 0)
            if name_item and name_item.text().startswith(file_name):
                rows_to_remove.append(row)
        
        # Remove rows from bottom to top to avoid index shifting
        for row in reversed(rows_to_remove):
            self.table_widget.removeRow(row)
        
        QMessageBox.information(self, "Bilgi", f"{file_name} dosyası ve verileri kaldırıldı.")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("DataQuasar")
        self.setGeometry(100, 100, 1200, 800)
        self.setStyleSheet(STYLE_SHEET)
        
        # Set window icon
        self.setWindowIcon(QIcon("icon/logo.png"))
        
        # Create main widget and layout
        main_widget = QWidget()
        main_layout = QVBoxLayout(main_widget)
        
        # Create horizontal layout for header
        header_layout = QHBoxLayout()
        
        # Logo and title
        logo_label = QLabel()
        logo_pixmap = QPixmap("icon/logo.png").scaled(60, 60, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        logo_label.setPixmap(logo_pixmap)
        
        # Modern, stylish title
        title_label = QLabel("DataQuasar")
        title_label.setStyleSheet("""
            font-size: 24px;
            font-weight: bold;
            color: #2c3e50;
            letter-spacing: 1px;
            text-shadow: 1px 1px 2px rgba(0,0,0,0.1);
        """)
        
        # Yardım butonu
        self.help_button = QPushButton('YARDIM')        
        self.help_button.setIcon(QIcon("icon/help.png"))
        self.help_button.clicked.connect(self.open_help)
        
        # Style the help button
        self.help_button.setStyleSheet("""
            QPushButton {
                background-color: #FFA500;
                color: white;
                font-weight: bold;
                padding: 6px 18px;
                border-radius: 6px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #FF8C00;
            }
        """)
        
        # Create horizontal layout for logo and title
        logo_title_layout = QHBoxLayout()
        logo_title_layout.addWidget(logo_label)
        logo_title_layout.addWidget(title_label)
        logo_title_layout.addStretch(1)
        
        # Header layout
        header_layout.addLayout(logo_title_layout)
        header_layout.addStretch(1)
        header_layout.addWidget(self.help_button)
        
        # Add header layout to main layout
        main_layout.addLayout(header_layout)
        
        # Create tab widget
        self.tab_widget = QTabWidget()
        
        # Create tabs
        self.scraper_tab = ScraperTab()
        self.excel_viewer_tab = ExcelViewer()
        
        # Add tabs to widget
        self.tab_widget.addTab(self.scraper_tab, "Veri Çekme")
        self.tab_widget.addTab(self.excel_viewer_tab, "Excel ve WhatsApp")
        
        # Add tab widget to main layout
        main_layout.addWidget(self.tab_widget)
        
        # Set the main widget as central widget
        self.setCentralWidget(main_widget)

    def open_help(self):
        help_url = "https://www.instagram.com/3.14quasar/"
        webbrowser.open(help_url)

class ScraperTab(QWidget):
    def __init__(self):
        super().__init__()
        self.thread = VeriÇekmeThread()
        self.thread.sinyal_guncelle.connect(self.durumu_guncelle)
        self.thread.sinyal_tamamlandi.connect(self.veri_cekmesi_tamamlandi)
        self.arama_terimleri = []
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout(self)
        
        # Arama bölümü
        search_frame = QFrame()
        search_frame.setObjectName("searchFrame")
        search_layout = QVBoxLayout(search_frame)
        
        # Arama başlığı
        header = QLabel("Arama Kriterleri")
        header.setStyleSheet("font-size: 16px; font-weight: bold; color: #17a2b8; margin-bottom: 10px;")
        search_layout.addWidget(header)
        
        # Arama formu
        form_layout = QHBoxLayout()
        
        self.arama_kutusu = QLineEdit()
        self.arama_kutusu.setPlaceholderText("Arama terimi girin")
        
        self.ekle_buton = QPushButton("Ekle")
        self.ekle_buton.setIcon(QIcon("icon/add.png"))
        self.ekle_buton.setCursor(QCursor(Qt.PointingHandCursor))
        self.ekle_buton.clicked.connect(self.arama_ekle)  # Bağlantı eklendi
        
        self.sil_buton = QPushButton("Sil")
        self.sil_buton.setIcon(QIcon("icon/delete.png"))
        self.sil_buton.setObjectName("deleteButton")
        self.sil_buton.setCursor(QCursor(Qt.PointingHandCursor))
        self.sil_buton.clicked.connect(self.arama_sil)  # Bağlantı eklendi
        
        form_layout.addWidget(self.arama_kutusu)
        form_layout.addWidget(self.ekle_buton)
        form_layout.addWidget(self.sil_buton)
        
        search_layout.addLayout(form_layout)
        
        # Arama tablosu eklendi
        self.arama_tablosu = QTableWidget()
        self.arama_tablosu.setColumnCount(1)
        self.arama_tablosu.setHorizontalHeaderLabels(["Arama Terimleri"])
        self.arama_tablosu.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        search_layout.addWidget(self.arama_tablosu)
        
        layout.addWidget(search_frame)

        # Veri sınırı kutusu eklendi
        limit_layout = QHBoxLayout()
        limit_label = QLabel("Veri Sınırı:")
        self.sınır_kutusu = QLineEdit()
        self.sınır_kutusu.setPlaceholderText("Opsiyonel - Maksimum veri sayısı")
        limit_layout.addWidget(limit_label)
        limit_layout.addWidget(self.sınır_kutusu)
        layout.addLayout(limit_layout)

        # Sonuç tablosu
        self.tablo = QTableWidget()
        self.tablo.setColumnCount(7)
        self.tablo.setHorizontalHeaderLabels([
            "İsim", "Adres", "Website", "Telefon", 
             "Ortalama Puan", "Enlem", "Boylam"
        ])
        self.tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.tablo)
        
        # Kontrol butonları
        control_layout = QHBoxLayout()
        
        self.baslat_buton = QPushButton("Veri Çekmeyi Başlat")
        self.baslat_buton.setIcon(QIcon("icon/start.png"))
        self.baslat_buton.setCursor(QCursor(Qt.PointingHandCursor))
        self.baslat_buton.clicked.connect(self.veri_cekmeyi_baslat)  # Bağlantı eklendi
        
        self.temizle_buton = QPushButton("Tabloyu Temizle")
        self.temizle_buton.setIcon(QIcon("icon/clear.png"))
        self.temizle_buton.setCursor(QCursor(Qt.PointingHandCursor))
        self.temizle_buton.clicked.connect(self.tabloyu_temizle)  # Bağlantı eklendi
        
        control_layout.addWidget(self.baslat_buton)
        control_layout.addWidget(self.temizle_buton)
        layout.addLayout(control_layout)
        
        # Durum çubuğu
        self.durum_etiketi = QLabel("Hazır")
        self.durum_etiketi.setStyleSheet("""
            padding: 8px;
            background: #e4f7f5;
            border-radius: 4px;
            color: #17a2b8;
        """)
        layout.addWidget(self.durum_etiketi)
        
    def arama_ekle(self):
        arama_terimi = self.arama_kutusu.text().strip()
        if not arama_terimi:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir arama terimi girin!")
            return
        
        if arama_terimi in self.arama_terimleri:
            QMessageBox.warning(self, "Uyarı", "Bu terim zaten eklendi!")
            return

        self.arama_terimleri.append(arama_terimi)
        self.guncelle_arama_tablosu()
        self.arama_kutusu.clear()

    def arama_sil(self):
        secili_satir = self.arama_tablosu.currentRow()
        if secili_satir < 0:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek için bir terim seçin!")
            return
        
        silinen_terim = self.arama_terimleri.pop(secili_satir)
        self.guncelle_arama_tablosu()
        QMessageBox.information(self, "Bilgi", f"'{silinen_terim}' silindi!")

    def guncelle_arama_tablosu(self):
        self.arama_tablosu.setRowCount(len(self.arama_terimleri))
        for satir, terim in enumerate(self.arama_terimleri):
            self.arama_tablosu.setItem(satir, 0, QTableWidgetItem(terim))

    def veri_cekmeyi_baslat(self):
        if not self.arama_terimleri:
            QMessageBox.warning(self, "Uyarı", "Lütfen en az bir arama terimi ekleyin!")
            return

        try:
            veri_siniri = int(self.sınır_kutusu.text().strip()) if self.sınır_kutusu.text().strip() else None
            self.thread.veri_sınırı = veri_siniri
        except ValueError:
            QMessageBox.warning(self, "Uyarı", "Lütfen geçerli bir sayı girin!")
            return

        self.baslat_buton.setEnabled(False)
        self.durum_etiketi.setText("Veri çekme işlemi başlatılıyor...")
        self.thread.arama_ayarla(self.arama_terimleri)
        self.thread.start()

    def durumu_guncelle(self, bilgi):
        self.durum_etiketi.setText(f"İşleniyor: {bilgi[0]} - {bilgi[1]} sonuç bulundu")

    def veri_cekmesi_tamamlandi(self, isletmeler):
        self.tablo.setRowCount(len(isletmeler))
        
        for satir, isletme in enumerate(isletmeler):
            self.tablo.setItem(satir, 0, QTableWidgetItem(str(isletme.isim)))
            self.tablo.setItem(satir, 1, QTableWidgetItem(str(isletme.adres)))
            self.tablo.setItem(satir, 2, QTableWidgetItem(str(isletme.website)))
            self.tablo.setItem(satir, 3, QTableWidgetItem(str(isletme.telefon)))            
            self.tablo.setItem(satir, 4, QTableWidgetItem(str(isletme.ortalama_puan)))
            self.tablo.setItem(satir, 5, QTableWidgetItem(str(isletme.enlem)))
            self.tablo.setItem(satir, 6, QTableWidgetItem(str(isletme.boylam)))

        self.baslat_buton.setEnabled(True)
        self.durum_etiketi.setText("Veri çekme işlemi tamamlandı!")
        QMessageBox.information(self, "Bilgi", "Veri çekme işlemi tamamlandı ve veriler Excel'e kaydedildi!")

    def tabloyu_temizle(self):
        self.tablo.clearContents()
        self.tablo.setRowCount(0)
        self.durum_etiketi.setText("Tablo temizlendi!") 

    def arama_ekle(self):
        arama_terimi = self.arama_kutusu.text().strip()
        if not arama_terimi:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir arama terimi girin!")
            return
        
        if arama_terimi in self.arama_terimleri:
            QMessageBox.warning(self, "Uyarı", "Bu terim zaten eklendi!")
            return

        self.arama_terimleri.append(arama_terimi)
        self.guncelle_arama_tablosu()
        self.arama_kutusu.clear()

    def arama_sil(self):
        secili_satir = self.arama_tablosu.currentRow()
        if secili_satir < 0:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek için bir terim seçin!")
            return
        
        silinen_terim = self.arama_terimleri.pop(secili_satir)
        self.guncelle_arama_tablosu()
        QMessageBox.information(self, "Bilgi", f"'{silinen_terim}' silindi!")

    def guncelle_arama_tablosu(self):
        self.arama_tablosu.setRowCount(len(self.arama_terimleri))
        for satir, terim in enumerate(self.arama_terimleri):
            self.arama_tablosu.setItem(satir, 0, QTableWidgetItem(terim))

    def veri_cekmeyi_baslat(self):
        if not self.arama_terimleri:
            QMessageBox.warning(self, "Uyarı", "Lütfen en az bir arama terimi ekleyin!")
            return

        try:
            # Kullanıcı sınır girdi mi kontrol et
            self.veri_sınırı = int(self.sınır_kutusu.text().strip()) if self.sınır_kutusu.text().strip() else None
        except ValueError:
            QMessageBox.warning(self, "Uyarı", "Lütfen geçerli bir sayı girin!")
            return

        self.baslat_buton.setEnabled(False)
        self.durum_etiketi.setText("Veri çekme işlemi başlatılıyor...")

        self.thread.arama_ayarla(self.arama_terimleri)
        self.thread.veri_sınırı = self.veri_sınırı  # Thread'e veri sınırını gönder
        self.thread.start()

    def durumu_guncelle(self, bilgi):
        self.durum_etiketi.setText(f"İşleniyor: {bilgi[0]} - {bilgi[1]} sonuç bulundu")

    def veri_cekmesi_tamamlandi(self, isletmeler):
        self.tablo.setRowCount(len(isletmeler))
        
        for satir, isletme in enumerate(isletmeler):
            self.tablo.setItem(satir, 0, QTableWidgetItem(str(isletme.isim)))
            self.tablo.setItem(satir, 1, QTableWidgetItem(str(isletme.adres)))
            self.tablo.setItem(satir, 2, QTableWidgetItem(str(isletme.website)))
            self.tablo.setItem(satir, 3, QTableWidgetItem(str(isletme.telefon)))            
            self.tablo.setItem(satir, 4, QTableWidgetItem(str(isletme.ortalama_puan)))
            self.tablo.setItem(satir, 5, QTableWidgetItem(str(isletme.enlem)))
            self.tablo.setItem(satir, 6, QTableWidgetItem(str(isletme.boylam)))

        self.baslat_buton.setEnabled(True)
        self.durum_etiketi.setText("Veri çekme işlemi tamamlandı!")
        QMessageBox.information(self, "Bilgi", "Veri çekme işlemi tamamlandı ve veriler Excel'e kaydedildi!")



if __name__ == '__main__':
    app = QApplication(sys.argv)
    
    # Font ayarları
    app.setFont(QFont('Segoe UI', 10))
    
    # Pencere stili
    app.setStyle('Fusion')
    
    # Tema renkleri
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor('#ffffff'))
    palette.setColor(QPalette.WindowText, QColor('#333333'))
    palette.setColor(QPalette.Base, QColor('#ffffff'))
    palette.setColor(QPalette.AlternateBase, QColor('#e4f7f5'))
    palette.setColor(QPalette.Text, QColor('#333333'))
    palette.setColor(QPalette.Button, QColor('#17a2b8'))
    palette.setColor(QPalette.ButtonText, QColor('#ffffff'))
    palette.setColor(QPalette.Highlight, QColor('#17a2b8'))
    palette.setColor(QPalette.HighlightedText, QColor('#ffffff'))
    app.setPalette(palette)

    # Login dialog'ı başlat
    dialog = LoginDialog()  # Doğrulama penceresini oluştur
    if dialog.exec_() == QDialog.Accepted:  # Eğer giriş başarılıysa
        window = MainWindow()  # Ana pencereyi başlat
        window.show()  # Ana pencereyi göster
    
    sys.exit(app.exec_())  # Uygulama çalıştır
